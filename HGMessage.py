from openpyxl import Workbook, load_workbook
import time
import os
import sys

today = time.strftime('%Y.%m.%d.') 
hours = int(time.strftime('%H'))
read_xlsx = load_workbook(r'source.xlsx',data_only= True)
sheet = read_xlsx.get_sheet_by_name("internet")
sheet2 = read_xlsx.get_sheet_by_name("paper")
settings = read_xlsx.get_sheet_by_name("settings")
settings_time = settings.cell(row=1, column=2) #시간(시) 설정 column

news_number = 0
team = "-문화홍보과-"

#보도시간 자동설정, 정오이전 실행시 09시 설정
if hours <= 12:
   settings_time.value = "09"
else :
   settings_time.value = '17' 
report_time = str(settings['B1'].value)

#신문/방송 뉴스 정리
def sortPaper ():
    paper = sheet2['A']
    paperName = sheet2
    category = "[신문/방송]"
    sorting(paper, paperName, category)
    
#인터넷 뉴스 정리
def sortInternet ():
    internet = sheet['A']
    internetName = sheet
    category = "[인터넷]"
    sorting(internet, internetName, category)

def sorting(sheetCol, sheetName, category):
    global news_number
    rows = []
    col = []
    internet = []
    check = 0
    for cell in sheetCol:
      if cell.value != None: #NoneType값은 append에서 제외됨
        col.append(cell.value) #col 값 저장
    i = 1
    for i in range(len(col)):
        i = i+1
        row = sheetName[i]
        for cell in row:
            if cell.value != None: #NoneType값은 append에서 제외됨
                rows.append(cell.value) #행 데이터 rows에 저장
            if rows:
                if check == 0:
                    print (category)
                    check = check + 1
        internet.append(str(rows[2]+'('+rows[1]+'_'+rows[3]+')\n'+rows[4]+'\n'))
        internet[i-1] = internet[i-1].replace(u'\xa0',u' ') #\xa0 에러 방지
        rows = []
        news_number = news_number + 1
        print(str(news_number)+'.'+internet[i-1])
    if news_number == 0:
        print("오늘 보도사항은 없습니다.\n")
        news_number += 1
        
if __name__ == "__main__":
    print("\n잠시만 기다려주세요..\n")
    sys.stdout = open('temp.txt','w') #결과값 txt파일로 내보내기       
    print("금일("+today+") "+report_time+"시까지 한강관련주요 보도사항입니다.\n")
    sortPaper()
    sortInternet()
    print(team) 
    os.startfile('temp.txt')
